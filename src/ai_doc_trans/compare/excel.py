from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Optional

import openpyxl


class CompareStatus(str, Enum):
    OK = "OK"
    MISSING = "MISSING"       # target cell is empty or identical to source
    STRUCTURE_DIFF = "STRUCTURE_DIFF"  # target sheet/cell does not exist


@dataclass
class CompareResult:
    position: str       # e.g. "Sheet1!A1"
    source_text: str
    target_text: Optional[str]
    status: CompareStatus


def _is_text_cell(cell) -> bool:
    if cell.value is None:
        return False
    if cell.data_type == "s" or isinstance(cell.value, str):
        return bool(str(cell.value).strip())
    return False


def _parse_position(position: str) -> tuple[str, str]:
    """Parse 'SheetName!A1' into (sheet_name, coordinate)."""
    sheet, _, coord = position.partition("!")
    return sheet, coord


def compare_excel(
    source_path: Path,
    target_path: Path,
) -> list[CompareResult]:
    """
    Compare source and target Excel files cell-by-cell using position info.

    For each text cell in source, finds the matching cell in target by
    sheet name and coordinate, then reports status:
      OK             – target has different (translated) content
      MISSING        – target cell is empty or identical to source
      STRUCTURE_DIFF – target sheet or cell coordinate doesn't exist
    """
    src_wb = openpyxl.load_workbook(str(source_path), data_only=True)
    tgt_wb = openpyxl.load_workbook(str(target_path), data_only=True)

    tgt_sheets = {ws.title: ws for ws in tgt_wb.worksheets}
    results: list[CompareResult] = []

    for src_sheet in src_wb.worksheets:
        for row in src_sheet.iter_rows():
            for cell in row:
                if not _is_text_cell(cell):
                    continue
                source_text = str(cell.value).strip()
                position = f"{src_sheet.title}!{cell.coordinate}"
                sheet_name, coordinate = _parse_position(position)

                if sheet_name not in tgt_sheets:
                    results.append(CompareResult(
                        position=position,
                        source_text=source_text,
                        target_text=None,
                        status=CompareStatus.STRUCTURE_DIFF,
                    ))
                    continue

                tgt_cell = tgt_sheets[sheet_name][coordinate]
                tgt_value = (
                    str(tgt_cell.value).strip()
                    if tgt_cell.value is not None
                    else ""
                )

                if not tgt_value or tgt_value == source_text:
                    status = CompareStatus.MISSING
                else:
                    status = CompareStatus.OK

                results.append(CompareResult(
                    position=position,
                    source_text=source_text,
                    target_text=tgt_value or None,
                    status=status,
                ))

    return results


def format_report(results: list[CompareResult]) -> str:
    total = len(results)
    ok = sum(1 for r in results if r.status == CompareStatus.OK)
    missing = sum(1 for r in results if r.status == CompareStatus.MISSING)
    struct = sum(1 for r in results if r.status == CompareStatus.STRUCTURE_DIFF)

    lines = [
        f"Compare report: {total} cells checked",
        f"  OK:              {ok}",
        f"  MISSING:         {missing}",
        f"  STRUCTURE_DIFF:  {struct}",
        "",
    ]
    for r in results:
        if r.status != CompareStatus.OK:
            lines.append(
                f"[{r.status.value:14s}] {r.position:20s} | source: {r.source_text!r}"
                + (f" | target: {r.target_text!r}" if r.target_text else "")
            )
    return "\n".join(lines)
