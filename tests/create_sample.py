"""Create a sample Excel file for testing."""
import openpyxl
from pathlib import Path

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

ws["A1"] = "Product Name"
ws["B1"] = "Description"
ws["C1"] = "Unit"

ws["A2"] = "NVL1000"       # product code → should be skipped
ws["B2"] = "Raw material"
ws["C2"] = "kg"

ws["A3"] = "NVL2000"       # product code → should be skipped
ws["B3"] = "Finished goods"
ws["C3"] = "pcs"

ws["A4"] = 42              # number → should be skipped
ws["B4"] = "Total weight"
ws["C4"] = "ton"

ws2 = wb.create_sheet("Info")
ws2["A1"] = "Document Title"
ws2["B1"] = "Translation Test"
ws2["A2"] = "Version"
ws2["B2"] = "1.0"          # number-like string, should be extracted

out = Path("tests/sample.xlsx")
out.parent.mkdir(exist_ok=True)
wb.save(str(out))
print(f"Created {out}")
